import pandas as pd
import openpyxl

path1 ="G:\\3 - Production Departments\\11- Scheduling\\Grinding\\Setup_Tracking.xlsx"
#path = "G:\\3 - Production Departments\\4 - Grinding\\0 - Department Documents\\4 - Programs & Software\\1 - Operating Software\\Setup Tracker\\Data\\setups_in_progress.csv"
path = ".//Data//setups_in_progress.csv"
sublist = [0,0]

def main():
    return readWIP()

def readwipSheet():
    wb = openpyxl.load_workbook(path1,read_only=True)
    sh = wb["Setups In Progress"]
    maxdatarow = 2 #initialize to first line, in case of empty list
    for row in sh.iter_rows(min_row=2, max_col=2, max_row=sh.max_row):
        for cell in row:
            if cell.value != None:
                maxdatarow = cell.row
    wiplist = [[None]*2 for i in range(maxdatarow-1)]
    for row in sh.iter_rows(min_row=2, max_col=2, max_row=maxdatarow):            
        for cell in row:
            if cell.value != None and cell.column == 1:
                wiplist[cell.row-2][0] = cell.value              
            if type(cell.value) == type("string") and cell.column == 2:
                wiplist[cell.row-2][1] = cell.value
    return wiplist

def readWIP():
    data = pd.read_csv(path)    
    wip = data.loc[:,'Job Number':'Machine'].values    
    return wip
    
    
        
if __name__ == "__main__":
    main()
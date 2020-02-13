import openpyxl

path = "G:\\Consolidated Training\\Factory Wide Training Matricies\\Grinding Matrix.xlsx"
emplist = []

def main():
    return readSheet()

def readSheet():
    wb = openpyxl.load_workbook(path)
    sh = wb["Sheet1"]
    for row in sh.iter_rows(min_row=7, max_col=1, max_row=sh.max_row, values_only=True):
        for cell in row:            
            if cell != None and "SHIFT" not in cell and "BELOW" not in cell:                
                emplist.append(cell)
            if "BELOW" in cell:
                return emplist     
    return emplist
if __name__ == "__main__":
    print(main())



    
import pandas as pd
import numpy as np
import openpyxl
import csv
from datetime import *

path = "G:\\3 - Production Departments\\11- Scheduling\\Grinding\\Setup_Tracking.xlsx"

#trackpath = "G:\\3 - Production Departments\\4 - Grinding\\0 - Department Documents\\4 - Programs & Software\\1 - Operating Software\\Setup Tracker\\Data\\tracked_setups.csv"
#untrackpath = "G:\\3 - Production Departments\\4 - Grinding\\0 - Department Documents\\4 - Programs & Software\\1 - Operating Software\\Setup Tracker\\Data\\untracked_setups.csv"

trackpath = ".\\Data\\tracked_setups.csv"
wippath = ".\\Data\\setups_in_progress.csv"
untrackpath = ".\\Data\\untracked_setups.csv"

path2 = "S:\\Metrics\\5 - Grinding\\setup_tracking.csv"

#will be obsolete
def openSetupBook():
    workbook = openpyxl.load_workbook(path)
    return workbook

def writeSetupCSV(jobn, mach, oper, tod, whe):
    with open(path2, 'a', newline= '') as setupcsv:
        tod = str(tod).split(' ')
        tod = tod[0]
        trackwriter = csv.writer(setupcsv, delimiter=',', quoting=csv.QUOTE_MINIMAL)        
        trackwriter.writerow([str(jobn)] + [str(mach)] + [str(oper)] + [str(tod)] + [str(whe)])
    
def writewipSheet(jobnum=0, machine='none', operator='none'):
    today = datetime.today().date() 
    #today = datetime.today().date() + timedelta(days=-1) #this line used for testing, subtracts 1 day 
    when = datetime.today().time()
    wb = openSetupBook()
    sh = wb["Setups In Progress"]
    openrow = 2
    for row in sh.iter_rows(min_row=2, max_row = sh.max_row, min_col=2, max_col=2):
        for cell in row:
            print(cell.row)
            if type(cell.value) == type('None'):
                openrow = cell.row + 1
    sh.cell(row=openrow,column = 1).value = jobnum
    sh.cell(row=openrow,column = 2).value = machine
    sh.cell(row=openrow,column = 3).value = operator
    sh.cell(row=openrow,column = 4).value = today
    sh.cell(row=openrow,column = 5).value = when    
    wb.save(path)

#writeWip saves to a CSV file instead of 
def writeWip(sd=['00000','nomach','no Opr']):
    now = pd.Timestamp.now()
    today = pd.Timestamp.date(now)
    when = pd.Timestamp.time(now)
    sd.extend([today,when])
    with open(wippath, 'a', newline= '\n') as wipcsv:
        trackwriter = csv.writer(wipcsv, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        trackwriter.writerow(sd)

    
def findCompletedWip(jobnum):
    wb = openSetupBook()
    wip_sh = wb["Setups In Progress"]
    cell_row = 1
    for row in wip_sh.iter_rows(min_row=2, max_row = wip_sh.max_row, max_col=1):
        for cell in row:
            if cell.value == jobnum:
                print("found it")
                cell_row = cell.row
                return cell_row
    return cell_row

def findCompWip(jobnum):
    df = pd.read_csv(wippath)
    dfilter = df['Job Number'].astype(str) == jobnum
    if df[dfilter].values.size > 0:
        found = df[dfilter].values        
        return found[0]
    else:        
        return None


def writeCompleted(jobnum="",comment=""):
    complete_data = findCompWip(jobnum)
    now = pd.Timestamp.now()
    today = pd.Timestamp.date(now)
    finish = pd.Timestamp.time(now)
    start = pd.to_datetime(complete_data[3] + 'T' + complete_data[4])    
    total_hours = pd.Timedelta(now - start).total_seconds()/3600  
    ml = complete_data.tolist() #ml is mylist
    ml.extend([finish,total_hours,comment])   
    
    with open(trackpath, 'a', newline='\n') as trackcsv:
        trackwriter = csv.writer(trackcsv, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        trackwriter.writerow(ml)
    
    
def writeCompletedSheet(jobnum = "", ls_reason = ""):    
    read_row = findCompletedWip(jobnum)
    if read_row == 1:
        print("job not found")
        return
    wb = openSetupBook()
    wip_sh = wb["Setups In Progress"]
    comp_sh = wb["Tracked Setups"]
    write_row = comp_sh.max_row+1
    end_date_time = datetime.today()
    end_date = end_date_time.date()    
    end_time = datetime.today().time().strftime("%H:%M:%S")
    start_date = wip_sh.cell(row=read_row, column=4).value
    start_time = wip_sh.cell(row=read_row, column=5).value        
    total_hours = calcHours(start_date, start_time, end_date_time)
    
    for row in wip_sh.iter_rows(min_row = read_row, max_row = read_row, max_col=5):
        for cell in row:
            comp_sh.cell(row = write_row, column= cell.column).value = cell.value
    jobnum = wip_sh.cell(row = read_row, column = 1).value
    machine = wip_sh.cell(row = read_row, column = 2).value
    operator = wip_sh.cell(row = read_row, column = 3).value
    today = wip_sh.cell(row = read_row, column = 4).value
    comp_sh.cell(row = write_row, column = 4).number_format = "mm/dd/yyyy"
    comp_sh.cell(row = write_row, column = 6).value = end_time
    comp_sh.cell(row = write_row, column = 7).value = total_hours
    comp_sh.cell(row = write_row, column = 7).number_format = "0.00"
    comp_sh.cell(row = write_row, column = 8).value = ls_reason
    wip_sh.delete_rows(read_row)
    wb.save(path)
    writeSetupCSV(jobnum, machine, operator, today, total_hours)
    
def calcHours(start, end_date_time):        
    dtime = end_date_time - start
    dhours = (dtime.days*24) + (dtime.seconds/3600)
    return dhours    
    
if __name__ == "__main__":    
    writeCompleted('12345')